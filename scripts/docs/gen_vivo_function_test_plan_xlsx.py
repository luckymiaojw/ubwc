#!/usr/bin/env python3
"""Generate UBWC ENC/DEC VIVO functional coverage test plan workbook."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "docs" / "ubwc_vivo_function_test_plan_cn.xlsx"
ENC_RTL = ROOT / "src" / "enc" / "ubwc_enc_vivo_top.sv"
DEC_RTL = ROOT / "src" / "dec" / "ubwc_dec_vivo_top.v"


def extract_inputs(path: Path) -> list[dict[str, str]]:
    ports: list[dict[str, str]] = []
    in_port_list = False
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.split("//", 1)[0].rstrip()
        if ")(" in line:
            in_port_list = True
            continue
        if not in_port_list:
            continue
        if line.strip() == ");":
            break
        m = re.match(
            r"\s*input\s+(?:wire|reg)?\s*(?P<width>\[[^\]]+\])?\s*(?P<name>[A-Za-z_][A-Za-z0-9_]*)\s*,?",
            line,
        )
        if not m:
            continue
        width = (m.group("width") or "1-bit").strip()
        width = re.sub(r"\s+", " ", width)
        ports.append({"name": m.group("name"), "width": width})
    return ports


COVERAGE_MODEL = [
    [
        "CVG-FORMAT",
        "ENC/DEC",
        "ENC: i_ci_format/i_rvi_format/o_co path；DEC: i_ci_format/i_cvi_format/o_rvo path",
        "RGBA8888; RGBA1010102; NV12_Y; NV12_UV; P010/G016_Y; P010/G016_UV",
        "0055-0058, 0061-0066",
        "对应 format 的 VIVO 端口 valid&&ready 至少命中一次；Y/UV plane 分开计数",
        "P0",
        "坐标不作为 coverpoint，只用于索引 golden",
    ],
    [
        "CVG-ALEN",
        "ENC/DEC",
        "ENC: o_co_alen 与 CVO payload；DEC: i_ci_alen 与 CVI/RVO payload",
        "alen=0,1,2,3,4,5,6,7",
        "metadata 覆盖集 + 0061-0066",
        "每个 alen bin 至少有一笔 tile 事务命中；payload beat 数与 alen 规则一致",
        "P0",
        "ENC 输入侧 i_ci_alen 当前默认 7；重点覆盖 VIVO 输出/DEC 输入 alen",
    ],
    [
        "CVG-META-FLAG",
        "DEC primary / ENC metadata compare",
        "DEC: i_ci_metadata[3:0] / o_dec_flag；ENC: written metadata low bits",
        "normal payload flag; padding/invalid flag=8; force-full flag=7; FC flag; SC flag=0",
        "0061-0066 + 含特殊 metadata vector",
        "每类 flag 至少命中一次，并检查 flag->alen/payload 行为",
        "P0",
        "DEC 的 i_ci_metadata 是 decoded meta_flag，不是 raw 8-bit metadata",
    ],
    [
        "CVG-DEC-SOLID-COLOR",
        "DEC",
        "raw metadata[7:6] != 0，经 decode 后进入 dec_vivo_top.i_ci_metadata",
        "SC raw[7:6]=01/10/11；flag=0；alen=0",
        "专用 SC vector 或扫描到有效 SC tile 的 vector",
        "SC tile 的 CI/CVI/RVO 数量和数据与 decoder golden 一致",
        "P0",
        "只针对有效 tile 区域；metadata padding 0 不算 SC/FC",
    ],
    [
        "CVG-DEC-FASTCLEAR",
        "DEC",
        "raw metadata[7:6]==0 && raw[4]==0",
        "FC raw[4]=0；flag={1,raw[3:1]}；alen=0",
        "专用 FC vector 或扫描到有效 FC tile 的 vector",
        "FC tile 不误取普通 payload；RVO 与 decoder golden 一致",
        "P0",
        "无效 metadata 区域的 0 不计入 FC 命中",
    ],
    [
        "CVG-CONSTANT-ALPHA",
        "DEC RGBA",
        "i_ci_alpha_mode[1:0]，raw metadata[5]/[0]",
        "alpha_mode=00/01/10/11；constant alpha; non-constant alpha",
        "RGBA8888/RGBA1010102 vector + 含 PCM alpha metadata 的 vector",
        "constant alpha 下 alen/payload 规则正确；non-constant alpha 走普通 PCM 规则",
        "P0",
        "重点是 DEC metadata decode 到 VIVO 的 alpha_mode，以及真实/假 VIVO 对 payload 的使用",
    ],
    [
        "CVG-RESOLUTION",
        "ENC/DEC",
        "APB active width/height, OTF active size, stored/pitch layout",
        "128x128; 256x160; 720x1548; 970x2134; 1888x1624; 3840x2016; 4096x600; max-width 4096",
        "已有 vector/cases + 新增 corner case",
        "每档分辨率至少 ENC/DEC 各跑一帧；非整 tile 尺寸不能卡死，padding 不进入有效输出",
        "P0",
        "覆盖分辨率类别，不覆盖每个坐标点",
    ],
    [
        "CVG-LOSSY",
        "ENC/DEC",
        "i_ci_lossy, is_lossy_rgba_2_1_format, metadata/payload",
        "lossless; NV12 lossy; RGBA8888 lossy 2:1",
        "0061,0062,0065,0066",
        "NV12 lossy 不走 RGBA2:1 地址分支；RGBA lossy 2:1 使用 y>>1 + odd row 128B",
        "P0",
        "DEC lossy 输出使用 decoder golden，不与 encoder 原始 linear bit-exact 比较",
    ],
]


ENC_TESTS = [
    ["ENC-FUNC-001", "format 覆盖", "CVG-FORMAT", "0055-0058,0061-0066", "跑 ENC fake/real smoke", "CI/RVI/CVO format 命中所有格式 bins；Y/UV plane 分别命中", "P0"],
    ["ENC-FUNC-002", "alen 与 CVO payload 长度", "CVG-ALEN", "metadata 覆盖集", "根据 metadata golden 生成 CO/CVO", "每个 alen bin 的 CVO beat/mask/last 与 golden 一致", "P0"],
    ["ENC-FUNC-003", "metadata low bits 写出", "CVG-META-FLAG", "0061-0066", "比较 AXI metadata dump", "只比较有效 metadata 规则定义的 bit field；padding 不误报特殊模式", "P0"],
    ["ENC-FUNC-004", "lossy RGBA2:1", "CVG-LOSSY", "0066", "配置 lossy=1 + RGBA2:1", "CVO 一 tile 一 payload；AXI AW 仅 4KB split；地址 y>>1 + odd row 128B", "P0"],
    ["ENC-FUNC-005", "NV12 lossy 普通地址", "CVG-LOSSY, CVG-FORMAT", "0062", "配置 lossy=1，NV12", "tile address 仍按 NV12 Y/UV layout，不走 RGBA2:1", "P0"],
    ["ENC-FUNC-006", "分辨率覆盖", "CVG-RESOLUTION", "128x128/256x160/720x1548/4096x600/3840x2016", "按 active size 送真实像素", "非整 tile 边界 mask 正确，frame done 后 idle", "P0"],
    ["ENC-FUNC-007", "backpressure 协议", "非功能 cover，协议检查", "任选两种格式", "随机 ready/backpressure", "valid&&ready 计数不丢不重，data/control stable", "P1"],
]


DEC_TESTS = [
    ["DEC-FUNC-001", "format 覆盖", "CVG-FORMAT", "0055-0058,0061-0066", "跑 DEC fake/real smoke", "CI/CVI/RVO format 命中所有格式 bins；Y/UV plane 分别命中", "P0"],
    ["DEC-FUNC-002", "alen decode 覆盖", "CVG-ALEN", "metadata 覆盖集", "AXI 读 metadata，经 decode 送 VIVO", "i_ci_alen 0..7 bins 命中；CVI/RVO payload 行为正确", "P0"],
    ["DEC-FUNC-003", "meta_flag decode 覆盖", "CVG-META-FLAG", "0061-0066 + 特殊 metadata vector", "扫描有效 metadata 并运行 DEC", "normal/padding/force-full/FC/SC flag 命中且行为正确", "P0"],
    ["DEC-FUNC-004", "solid color", "CVG-DEC-SOLID-COLOR", "含 SC 有效 tile 的 vector", "raw metadata[7:6]!=0", "flag=0、alen=0；RVO 与 decoder golden 一致", "P0"],
    ["DEC-FUNC-005", "fastclear", "CVG-DEC-FASTCLEAR", "含 FC 有效 tile 的 vector", "raw[7:6]==0 && raw[4]==0", "alen=0，不误取普通 payload；RVO 与 decoder golden 一致", "P0"],
    ["DEC-FUNC-006", "constant alpha", "CVG-CONSTANT-ALPHA", "RGBA metadata 覆盖集", "覆盖 alpha_mode[1:0]", "constant alpha 和 non-constant alpha 的 alen/payload 行为正确", "P0"],
    ["DEC-FUNC-007", "lossy decoder 输出", "CVG-LOSSY", "0062,0066 decoder golden", "用 compressed+metadata 解码", "RVO/OTF 与 decoder golden 比较，不与 encoder 原始 linear 比较", "P0"],
    ["DEC-FUNC-008", "分辨率覆盖", "CVG-RESOLUTION", "128x128/256x160/720x1548/4096x600/3840x2016", "按 active size 配置 OTF 输出", "输出有效宽高正确；最后一行后 FIFO/status idle", "P0"],
    ["DEC-FUNC-009", "backpressure 协议", "非功能 cover，协议检查", "任选两种格式", "随机 ready/AXI delay", "CI/CVI/RVO 不丢不重，last 不提前", "P1"],
]


RESOLUTION_BINS = [
    ["RES-001", "128x128", "small, tile aligned", "RGBA/NV12 任一", "smoke / SRAM basic"],
    ["RES-002", "256x160", "small-medium", "R7130 pattern", "vector_db 应包含"],
    ["RES-003", "720x1548", "non tile aligned height", "NV12/RGBA", "active height 与 stored height 区分"],
    ["RES-004", "970x2134", "non tile aligned width/height", "NV12/P010/RGBA", "padding/mask corner"],
    ["RES-005", "1888x1624", "mid-large non aligned", "NV12", "DEC tile_to_otf UV/Y 调度"],
    ["RES-006", "3840x2016", "large lossy RGBA", "0066", "RGBA lossy 2:1"],
    ["RES-007", "4096x600", "max width", "0061-0065", "COM_BUF_AW=12/max-width regression"],
]


PORT_USAGE_NOTE = {
    "i_clk": "协议/时序检查；不是功能 coverpoint",
    "i_reset": "reset 流程检查；不是功能 coverpoint",
    "i_sreset": "soft reset 流程检查；不是功能 coverpoint",
    "i_ubwc_en": "enable/idle 协议检查；不是功能 coverpoint",
    "i_ci_valid": "用于采样 format/alen/meta_flag/alpha/lossy coverpoint",
    "i_ci_input_type": "配置合法性检查；不是主要功能 coverpoint",
    "i_ci_alen": "alen coverpoint",
    "i_ci_format": "format coverpoint",
    "i_ci_metadata": "ENC: metadata low bits/checker；DEC: meta_flag coverpoint",
    "i_ci_forced_pcm": "PCM 行为检查；不单独作为本版 coverpoint",
    "i_ci_xcoord": "不覆盖坐标；仅用于 golden lookup/debug",
    "i_ci_ycoord": "不覆盖坐标；仅用于 golden lookup/debug",
    "i_ci_fcnt": "多帧/slot 协议检查；不是功能 coverpoint",
    "i_ci_lossy": "lossy coverpoint",
    "i_ci_alpha_mode": "constant alpha coverpoint",
    "i_ci_sb": "slot 协议检查；不是功能 coverpoint",
    "i_rvi_valid": "用于采样 uncompressed input；不是独立功能 coverpoint",
    "i_rvi_data": "数据一致性检查；不是 cover bin",
    "i_rvi_mask": "padding/mask 检查；与 resolution coverpoint 关联",
    "i_rvi_last": "协议检查；不是功能 coverpoint",
    "i_rvi_format": "format coverpoint",
    "i_rvi_xcoord": "不覆盖坐标；仅用于 golden lookup/debug",
    "i_rvi_ycoord": "不覆盖坐标；仅用于 golden lookup/debug",
    "i_rvi_fcnt": "多帧/slot 协议检查；不是功能 coverpoint",
    "i_cvi_valid": "用于采样 compressed input；不是独立功能 coverpoint",
    "i_cvi_data": "数据一致性检查；不是 cover bin",
    "i_cvi_last": "协议检查；不是功能 coverpoint",
    "i_cvi_format": "format coverpoint",
    "i_cvi_xcoord": "不覆盖坐标；仅用于 golden lookup/debug",
    "i_cvi_ycoord": "不覆盖坐标；仅用于 golden lookup/debug",
    "i_cvi_fcnt": "多帧/slot 协议检查；不是功能 coverpoint",
    "i_co_ready": "backpressure 协议检查；不是功能 coverpoint",
    "i_cvo_ready": "backpressure 协议检查；不是功能 coverpoint",
    "i_rvo_ready": "backpressure 协议检查；不是功能 coverpoint",
}


def port_note(name: str) -> str:
    if re.match(r"i_ci_ubwc_cfg_\d+$", name):
        return "UBWC cfg 配置合法性检查；不是本版主要功能 coverpoint"
    return PORT_USAGE_NOTE.get(name, "接口协议/稳定性检查；不是本版主要功能 coverpoint")


def port_coverpoints(name: str) -> str:
    note = port_note(name)
    cps = []
    if "format coverpoint" in note:
        cps.append("CVG-FORMAT")
    if "alen coverpoint" in note:
        cps.append("CVG-ALEN")
    if "meta_flag coverpoint" in note:
        cps.append("CVG-META-FLAG")
    if "constant alpha coverpoint" in note:
        cps.append("CVG-CONSTANT-ALPHA")
    if "lossy coverpoint" in note:
        cps.append("CVG-LOSSY")
    if "resolution coverpoint" in note:
        cps.append("CVG-RESOLUTION")
    return ", ".join(cps) if cps else "-"


def add_rows(ws, title: str, headers: list[str], rows: list[list[object]]) -> None:
    ws.append([title])
    ws.append(headers)
    for row in rows:
        ws.append(row)


def style_sheet(ws) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    p0_fill = PatternFill("solid", fgColor="FCE4D6")
    p1_fill = PatternFill("solid", fgColor="FFF2CC")
    p2_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = ws.max_column
    if ws.max_row >= 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        cell = ws.cell(1, 1)
        cell.fill = title_fill
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        pri = str(row[-1].value)
        if pri == "P0":
            row[-1].fill = p0_fill
        elif pri == "P1":
            row[-1].fill = p1_fill
        elif pri == "P2":
            row[-1].fill = p2_fill

    for idx, width in enumerate([18, 28, 36, 34, 34, 50, 12, 42, 24, 28], start=1):
        if idx <= max_col:
            ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    enc_ports = extract_inputs(ENC_RTL)
    dec_ports = extract_inputs(DEC_RTL)

    wb = Workbook()
    wb.remove(wb.active)

    overview = wb.create_sheet("Overview")
    overview_rows = [
        ["项目", "内容"],
        ["文档目的", "围绕 ubwc_enc_vivo_top / ubwc_dec_vivo_top 的功能覆盖和 directed test plan。"],
        ["覆盖口径", "本版 functional coverage 只覆盖 format、alen、meta_flag、DEC solid color、DEC fastclear、constant alpha、图像分辨率、lossy 类型。"],
        ["坐标口径", "xcoord/ycoord 不作为 coverpoint；只用于 golden lookup、地址计算、错误定位。"],
        ["端口口径", "VIVO TOP input 仍列入 Port Usage 表；但只有上述功能维度作为 coverpoint。"],
        ["ENC input 数量", len(enc_ports)],
        ["DEC input 数量", len(dec_ports)],
    ]
    add_rows(overview, "UBWC VIVO Function Coverage Plan Overview", overview_rows[0], overview_rows[1:])

    cov_ws = wb.create_sheet("Coverage Model")
    add_rows(cov_ws, "Functional Coverage Model", ["Coverpoint ID", "Scope", "采样信号/来源", "Bins", "推荐 case/vector", "Hit / Pass rule", "优先级", "备注"], COVERAGE_MODEL)

    enc_ws = wb.create_sheet("ENC Directed Tests")
    add_rows(enc_ws, "ENC VIVO Directed Tests", ["TC ID", "测试目标", "覆盖项", "推荐 case/vector", "激励方法", "通过标准", "优先级"], ENC_TESTS)

    dec_ws = wb.create_sheet("DEC Directed Tests")
    add_rows(dec_ws, "DEC VIVO Directed Tests", ["TC ID", "测试目标", "覆盖项", "推荐 case/vector", "激励方法", "通过标准", "优先级"], DEC_TESTS)

    res_ws = wb.create_sheet("Resolution Bins")
    add_rows(res_ws, "Resolution Coverage Bins", ["Bin ID", "分辨率", "类型", "推荐格式/case", "覆盖目的"], RESOLUTION_BINS)

    enc_port_rows = [["ENC", p["name"], p["width"], port_coverpoints(p["name"]), port_note(p["name"])] for p in enc_ports]
    enc_port_ws = wb.create_sheet("ENC Input Usage")
    add_rows(enc_port_ws, "ENC VIVO Input Usage", ["模块", "Input Port", "位宽", "关联 coverpoint", "说明"], enc_port_rows)

    dec_port_rows = [["DEC", p["name"], p["width"], port_coverpoints(p["name"]), port_note(p["name"])] for p in dec_ports]
    dec_port_ws = wb.create_sheet("DEC Input Usage")
    add_rows(dec_port_ws, "DEC VIVO Input Usage", ["模块", "Input Port", "位宽", "关联 coverpoint", "说明"], dec_port_rows)

    closure = wb.create_sheet("Closure")
    closure_rows = [
        ["项目", "关闭条件", "优先级"],
        ["Format", "所有 format bins 在 ENC/DEC 至少各命中一次；Y/UV plane 分开统计", "P0"],
        ["Alen", "alen=0..7 至少各命中一次；payload beat/mask/last 与规则一致", "P0"],
        ["Meta flag", "normal、padding/invalid、force-full、FC、SC flag 均命中", "P0"],
        ["DEC special modes", "solid color、fastclear、constant alpha 均有有效 tile 命中，不统计 padding 区域", "P0"],
        ["Resolution", "每个 resolution bin 至少 ENC/DEC 各跑一帧；非整 tile 不死锁", "P0"],
        ["Lossy", "NV12 lossy 和 RGBA8888 lossy 2:1 均覆盖；DEC 使用 decoder golden", "P0"],
        ["Coordinates", "不要求坐标覆盖率；仅在 first mismatch/debug dump 中记录坐标", "P1"],
    ]
    add_rows(closure, "Coverage Closure Criteria", closure_rows[0], closure_rows[1:])

    for ws in wb.worksheets:
        style_sheet(ws)
        if ws.title == "Overview":
            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 110
        if ws.title in {"ENC Input Usage", "DEC Input Usage"}:
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 26
            ws.column_dimensions["E"].width = 80

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
